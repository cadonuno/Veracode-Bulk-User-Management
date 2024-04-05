import sys
import requests
import getopt
import json
import urllib.parse
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
import openpyxl
import time
import xml.etree.ElementTree as ET  # for parsing XML

from veracode_api_signing.credentials import get_credentials

class NoExactMatchFoundException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message
    
class NoResultFoundException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message

class UnableToCreateTeamException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message

USERNAME_COLUMN = 1 
TEAMS_COLUMN = 2
ROLES_COLUMN = 3
TEAMS_MANAGED_COLUMN = 4
STATUS_SUCCESS = "success"
FIRST_ROW=3
LAST_COLUMN = TEAMS_MANAGED_COLUMN
TEAM_ADMIN_RELATIONSHIP = "ADMIN"
TEAM_MEMBER_RELATIONSHIP = "MEMBER"

teams_cache = {}

json_headers = {
    "User-Agent": "Bulk application creation - python script",
    "Content-Type": "application/json"
}

failed_attempts = 0
max_attempts_per_request = 10
sleep_time = 10


def print_help():
    """Prints command line options and exits"""
    print("""bulk-update-users-permissions.py -f <excel_file_with_user_information> [-d]"
        Reads all lines in <excel_file_with_user_information>, for each line, it will modify the user profile
        If a field is left empty, it will not be modified, to clear assigned teams, set the value to NONE (case sensitive). 
        If a team does not exist, it will be created.
""")
    sys.exit()

def row_as_number(row):
    return row-1

def request_encode(value_to_encode):
    return urllib.parse.quote(value_to_encode, safe='')

def find_exact_match(list, to_find, field_name):
    for index in range(len(list)):
        if list[index][field_name].lower() == to_find.lower():
            return list[index]
    print(f"Unable to find a member of list with '{field_name}' equal to '{to_find}'")
    raise NoExactMatchFoundException(f"Unable to find a member of list with {field_name} equal to {to_find}")

def get_item_from_api_call(api_base, api_to_call, item_to_find, list_name, field_to_check, field_to_get, is_exact_match, verbose):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    path = f"{api_base}{api_to_call}"
    if verbose:
        print(f"Calling: {path}")

    response = requests.get(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers)
    data = response.json()

    if response.status_code == 200:
        if verbose:
            print(data)
        if "_embedded" in data and len(data["_embedded"][list_name]) > 0:
            return (find_exact_match(data["_embedded"][list_name], item_to_find, field_to_check) if is_exact_match else data["_embedded"][list_name][0])[field_to_get]
        else:
            error_message = f"ERROR: No {list_name} named '{item_to_find}' found"
            print(error_message)
            raise NoResultFoundException(error_message)
    else:
        print(f"ERROR: trying to get {list_name} named {item_to_find}")
        print(f"ERROR: code: {response.status_code}")
        print(f"ERROR: value: {data}")
        failed_attempts+=1
        if (failed_attempts < max_attempts_per_request):
            time.sleep(sleep_time)
            return get_item_from_api_call(api_base, api_to_call, item_to_find, list_name, field_to_check, field_to_get, verbose)
        else:
            error_message = f"ERROR: trying to get {list_name} named {item_to_find}"
            print(error_message)
            raise NoResultFoundException(error_message)

def list_roles(roles):
    if not roles:
        return ""
    all_roles = roles.split(",")
    inner_role_list = ""
    for role_name in all_roles:
        inner_role_list = inner_role_list + (""",
            """ if inner_role_list else "") + f'{{ "role_name": "{role_name.strip()}" }}'
    if inner_role_list:
        return f'''
            "roles": [
                {inner_role_list}
            ]'''
    else:
        return None
    
def create_team_for_name(api_base, team_name, verbose):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    path = f"{api_base}api/authn/v2/teams"
    if verbose:
        print(f"Calling: {path}")

    request_content=f'''{{
            "team_name": "{team_name}"
        }}'''
    if verbose:
        print(request_content)

    response = requests.post(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, json=json.loads(request_content))

    if verbose:
        print(f"status code {response.status_code}")
        body = response.json()
        if body:
            print(body)

    body = response.json()
    if response.status_code == 201:
        print(f"Successfully created team: {team_name}.")
        return body["team_id"]
    else:
        if (body):
            raise UnableToCreateTeamException(f"Unable to create team: {response.status_code} - {body}")
        else:
            raise UnableToCreateTeamException(f"Unable to create team: {response.status_code}")

def get_team_id_from_name(api_base, team_name, verbose):
    try:
        team_guid = get_item_from_api_call(api_base, "api/authn/v2/teams?all_for_org=true&team_name="+ request_encode(team_name), team_name, "teams", "team_name", "team_id", True, verbose)
    except (NoExactMatchFoundException, NoResultFoundException):
        team_guid = None
        print(f"No team named {team_name} found. Creating new team.")
    if team_guid:
        return team_guid
    else:
        return create_team_for_name(api_base, team_name, verbose)
    
def get_all_teams_json(api_base, all_teams, all_teams_managed, verbose):
    global teams_cache
    all_teams_json = {}
    for untrimmed_name in all_teams:
        team_name = untrimmed_name.strip()
        team_id = None
        if team_name in teams_cache:
            team_id = teams_cache[team_name]
        else:
            team_id = get_team_id_from_name(api_base, team_name, verbose)
            teams_cache[team_name] = team_id
        if team_id:
            new_team = {}
            new_team["team_id"] = team_id
            new_team["relationship"] = TEAM_MEMBER_RELATIONSHIP
            all_teams_json[team_name] = new_team

    for untrimmed_name in all_teams_managed:
        team_name = untrimmed_name.strip()   
        team_id = None
        if team_name in all_teams_json:
            all_teams_json[team_name]["relationship"] = TEAM_ADMIN_RELATIONSHIP
        else:
            if team_name in teams_cache:
                team_id = teams_cache[team_name]
            else:
                team_id = get_team_id_from_name(api_base, team_name, verbose)
                teams_cache[team_name] = team_id
            if team_id:
                new_team = {}
                new_team["team_id"] = team_id
                new_team["relationship"] = TEAM_ADMIN_RELATIONSHIP
                all_teams_json[team_name] = new_team
    return all_teams_json if all_teams_json else None
            
def list_teams(api_base, teams, teamsManaged, verbose):
    if not teams and not teamsManaged:
        return ""
    if teams == "NONE":
        return '"teams": []'
    all_teams_json = get_all_teams_json(api_base, teams.split(",") if teams else {}, teamsManaged.split(",") if teamsManaged else {}, verbose)

    inner_team_list = ""
    for team in all_teams_json.values():
        team_value = f'''{{
            "team_id": "{team["team_id"]}",
            "relationship": {{
                "name": "{team["relationship"]}"
            }}
        }}'''
        inner_team_list = inner_team_list + (""",
        """ if inner_team_list else "") + team_value
    if inner_team_list:
        return f'''
            "teams": [
                {inner_team_list}
            ]'''
    else:
        return ""

def get_error_node_value(body):
    inner_node = ET.XML(body)
    if inner_node.tag == "error" and not inner_node == None:
        return inner_node.text
    else:
        return ""
    
def get_user_guid(api_base, username, verbose):
    return get_item_from_api_call(api_base, "api/authn/v2/users?deleted=false&user_name="+ request_encode(username.strip()), username.strip(), "users", "user_name", "user_id", True, verbose)

def modify_user(api_base, username, teams, roles, teamsManaged, verbose):
    if not username:
        error_message = "Empty user field found"
        print(error_message)
        return error_message
    if verbose:
        print(f"Updating user permissions for: {username}")

    user_guid = get_user_guid(api_base, username, verbose)

    if not user_guid:
        error_message = f"User with name '{username}' not found"
        print(error_message)
        return error_message
    
    path = f"{api_base}api/authn/v2/users/{user_guid}?partial=true"
    roles_json = list_roles(roles)
    teams_json = list_teams(api_base, teams, teamsManaged, verbose)
    if not roles_json and not teams_json:
        error_message = f"No teams or roles found for user '{username}', skipped this line"
        print(error_message)        
        return error_message
    
    content = roles_json
    if teams_json:
        if roles_json:
            content = content + """, 
            """ + teams_json
        else:
            content = teams_json
    request_content=f'''{{
            {content}
        }}'''
    if verbose:
        print(request_content)

    response = requests.put(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, json=json.loads(request_content))

    if verbose:
        print(f"status code {response.status_code}")
        body = response.json()
        if body:
            print(body)
    if response.status_code == 200:
        print(f"Successfully modified user permissions for {username}.")
        return STATUS_SUCCESS
    else:
        body = response.json()
        if (body):
            return f"Unable to modify user permissions: {response.status_code} - {body}"
        else:
            return f"Unable to modify user permissions: {response.status_code}"
    

def modify_all_users(api_base, file_name, verbose):
    global failed_attempts
    excel_file = openpyxl.load_workbook(file_name)
    excel_sheet = excel_file.active    
    try:
        for row in range(FIRST_ROW, excel_sheet.max_row+1):
            failed_attempts = 0
            status=excel_sheet.cell(row = row, column = LAST_COLUMN+1).value
            if (status == STATUS_SUCCESS):
                print(f"Skipping row {row-FIRST_ROW+1} as it was already done (physical row: {row})")
            else:
                try:
                    print(f"Importing row {row-FIRST_ROW+1}/{excel_sheet.max_row-FIRST_ROW+1} (physical row: {row}):")
                    status = modify_user(api_base, 
                                         excel_sheet.cell(row = row, column = USERNAME_COLUMN).value, 
                                         excel_sheet.cell(row = row, column = TEAMS_COLUMN).value, 
                                         excel_sheet.cell(row = row, column = ROLES_COLUMN).value, 
                                         excel_sheet.cell(row = row, column = TEAMS_MANAGED_COLUMN).value, 
                                         verbose)
                    print(f"Finished importing row {row-FIRST_ROW+1}/{excel_sheet.max_row-FIRST_ROW+1} (physical row: {row})")
                    print("---------------------------------------------------------------------------")
                except (NoExactMatchFoundException, UnableToCreateTeamException, NoResultFoundException) as e:
                    status= e.get_message()
                excel_sheet.cell(row = row, column = LAST_COLUMN+1).value=status
    finally:
        excel_file.save(filename=file_name)

def get_api_base():
    api_key_id, api_key_secret = get_credentials()
    api_base = "https://api.veracode.{instance}/"
    if api_key_id.startswith("vera01"):
        return api_base.replace("{instance}", "eu", 1)
    else:
        return api_base.replace("{instance}", "com", 1)

def main(argv):
    """Allows for bulk modifications for user permissions"""
    global failed_attempts
    excel_file = None
    try:
        verbose = False
        file_name = ''

        opts, args = getopt.getopt(argv, "hdf:", ["file_name="])
        for opt, arg in opts:
            if opt == '-h':
                print_help()
            if opt == '-d':
                verbose = True
            if opt in ('-f', '--file_name'):
                file_name=arg

        api_base = get_api_base()
        if file_name:
            modify_all_users(api_base, file_name, verbose)
        else:
            print_help()
    except requests.RequestException as e:
        print("An error occurred!")
        print(e)
        sys.exit(1)
    finally:
        if excel_file:
            excel_file.save(filename=file_name)


if __name__ == "__main__":
    main(sys.argv[1:])
