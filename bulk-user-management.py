import sys
import requests
import getopt
import json
import urllib.parse
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
import openpyxl
import time
import xml.etree.ElementTree as ET  # for parsing XML

from veracode_api_signing.credentials import get_credentials


class NoExactMatchFoundException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message
    
class NoResultFoundException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message

class UnableToCreateTeamException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message


API_SERVICE_ACCOUNT_COLUMN = 1

ACTIVE_COLUMN = 2
USERNAME_COLUMN = 3
FIRST_NAME_COLUMN = 4
LAST_NAME_COLUMN = 5

EMAIL_COLUMN = 6
PHONE_COLUMN = 7

POSITION_COLUMN = 8
RESTRICT_LOGIN_IPS_COLUMN = 9
LOGIN_ENABLED_COLUMN = 10

CUSTOM_1_COLUMN = 11
CUSTOM_2_COLUMN = 12
CUSTOM_3_COLUMN = 13
CUSTOM_4_COLUMN = 14
CUSTOM_5_COLUMN = 15

TEAMS_COLUMN = 16
ROLES_COLUMN = 17
TEAMS_MANAGED_COLUMN = 18

FIRST_ROW=3
LAST_COLUMN = TEAMS_MANAGED_COLUMN
STATUS_COLUMN = LAST_COLUMN+1
API_ID_COLUMN = STATUS_COLUMN+1
API_SECRET_COLUMN = API_ID_COLUMN+1
STATUS_SUCCESS = "success"
TEAM_ADMIN_RELATIONSHIP = "ADMIN"
TEAM_MEMBER_RELATIONSHIP = "MEMBER"
NONE = "NONE"

verify_ssl = True

teams_cache = {}

json_headers = {
    "User-Agent": "Bulk application creation - python script",
    "Content-Type": "application/json"
}

failed_attempts = 0
max_attempts_per_request = 10
sleep_time = 10


def print_help():
    """Prints command line options and exits"""
    print("""bulk-user-management.py -f <excel_file_with_user_information> [-c] [-g] [-d]"
        Reads all lines in <excel_file_with_user_information>, for each line, it will modify the user profile
        If a field is left empty, it will not be modified, to clear assigned teams, set the value to NONE (case sensitive). 
        If a team does not exist, it will be created.
        To create new users, you can pass the -c flag.
        You can use the -g flag to generate API credentials for new API accounts.
""")
    sys.exit()

def row_as_number(row):
    return row-1

def request_encode(value_to_encode):
    return urllib.parse.quote(value_to_encode, safe='')

def find_exact_match(list, to_find, field_name):
    for index in range(len(list)):
        if list[index][field_name].lower() == to_find.lower():
            return list[index]
    print(f"Unable to find a member of list with '{field_name}' equal to '{to_find}'")
    raise NoExactMatchFoundException(f"Unable to find a member of list with {field_name} equal to {to_find}")

def get_item_from_api_call(api_base, api_to_call, item_to_find, list_name, field_to_check, field_to_get, is_exact_match, verbose, error_on_not_found=True):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    path = f"{api_base}{api_to_call}"
    if verbose:
        print(f"Calling: {path}")

    response = requests.get(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, verify=verify_ssl)
    data = response.json()

    if response.status_code == 200:
        if verbose:
            print(data)
        if "_embedded" in data and len(data["_embedded"][list_name]) > 0:
            return (find_exact_match(data["_embedded"][list_name], item_to_find, field_to_check) if is_exact_match else data["_embedded"][list_name][0])[field_to_get]
        elif error_on_not_found:
            error_message = f"ERROR: No {list_name} named '{item_to_find}' found"
            print(error_message)
            raise NoResultFoundException(error_message)
        else:
            return ""
    else:
        print(f"ERROR: trying to get {list_name} named {item_to_find}")
        print(f"ERROR: code: {response.status_code}")
        print(f"ERROR: value: {data}")
        failed_attempts+=1
        if (failed_attempts < max_attempts_per_request):
            time.sleep(sleep_time)
            return get_item_from_api_call(api_base, api_to_call, item_to_find, list_name, field_to_check, field_to_get, is_exact_match, verbose, error_on_not_found)
        else:
            error_message = f"ERROR: trying to get {list_name} named {item_to_find}"
            print(error_message)
            raise NoResultFoundException(error_message)

def list_roles(roles):
    if not roles:
        return ""
    all_roles = roles.split(",")
    inner_role_list = ""
    for role_name in all_roles:
        inner_role_list = inner_role_list + (""",
            """ if inner_role_list else "") + f'{{ "role_name": "{role_name.strip()}" }}'
    if inner_role_list:
        return f'''
            "roles": [
                {inner_role_list}
            ]'''
    else:
        return None
    
def create_team_for_name(api_base, team_name, verbose):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    path = f"{api_base}api/authn/v2/teams"
    if verbose:
        print(f"Calling: {path}")

    request_content=f'''{{
            "team_name": "{team_name}"
        }}'''
    if verbose:
        print(request_content)

    response = requests.post(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, json=json.loads(request_content), verify=verify_ssl)

    if verbose:
        print(f"status code {response.status_code}")
        body = response.json()
        if body:
            print(body)

    body = response.json()
    if response.status_code == 201:
        print(f"Successfully created team: {team_name}.")
        return body["team_id"]
    else:
        if (body):
            raise UnableToCreateTeamException(f"Unable to create team: {response.status_code} - {body}")
        else:
            raise UnableToCreateTeamException(f"Unable to create team: {response.status_code}")

def get_team_id_from_name(api_base, team_name, verbose):
    try:
        team_guid = get_item_from_api_call(api_base, "api/authn/v2/teams?all_for_org=true&team_name="+ request_encode(team_name), team_name, "teams", "team_name", "team_id", True, verbose)
    except (NoExactMatchFoundException, NoResultFoundException):
        team_guid = None
        print(f"No team named {team_name} found. Creating new team.")
    if team_guid:
        return team_guid
    else:
        return create_team_for_name(api_base, team_name, verbose)
    
def get_all_teams_json(api_base, all_teams, all_teams_managed, verbose):
    global teams_cache
    all_teams_json = {}
    for untrimmed_name in all_teams:
        team_name = untrimmed_name.strip()
        team_id = None
        if team_name in teams_cache:
            team_id = teams_cache[team_name]
        else:
            team_id = get_team_id_from_name(api_base, team_name, verbose)
            teams_cache[team_name] = team_id
        if team_id:
            new_team = {}
            new_team["team_id"] = team_id
            new_team["relationship"] = TEAM_MEMBER_RELATIONSHIP
            all_teams_json[team_name] = new_team

    for untrimmed_name in all_teams_managed:
        team_name = untrimmed_name.strip()   
        team_id = None
        if team_name in all_teams_json:
            all_teams_json[team_name]["relationship"] = TEAM_ADMIN_RELATIONSHIP
        else:
            if team_name in teams_cache:
                team_id = teams_cache[team_name]
            else:
                team_id = get_team_id_from_name(api_base, team_name, verbose)
                teams_cache[team_name] = team_id
            if team_id:
                new_team = {}
                new_team["team_id"] = team_id
                new_team["relationship"] = TEAM_ADMIN_RELATIONSHIP
                all_teams_json[team_name] = new_team
    return all_teams_json if all_teams_json else None
            
def list_teams(api_base, teams, teamsManaged, verbose):
    if not teams and not teamsManaged:
        return ""
    if teams == NONE:
        return '"teams": []'
    all_teams_json = get_all_teams_json(api_base, teams.split(",") if teams else {}, teamsManaged.split(",") if teamsManaged else {}, verbose)

    inner_team_list = ""
    for team in all_teams_json.values():
        team_value = f'''{{
            "team_id": "{team["team_id"]}",
            "relationship": {{
                "name": "{team["relationship"]}"
            }}
        }}'''
        inner_team_list = inner_team_list + (""",
        """ if inner_team_list else "") + team_value
    if inner_team_list:
        return f'''
            "teams": [
                {inner_team_list}
            ]'''
    else:
        return ""

def get_error_node_value(body):
    inner_node = ET.XML(body)
    if inner_node.tag == "error" and not inner_node == None:
        return inner_node.text
    else:
        return ""
    
def get_user_guid(api_base, username, verbose):
    try:
        return get_item_from_api_call(api_base, "api/authn/v2/users?deleted=false&user_name="+ request_encode(username.strip()), username.strip(), "users", "user_name", "user_id", True, verbose, False)
    except (NoResultFoundException, NoExactMatchFoundException):
        print(f"Active user {username} not found, looking for inactive users")
        return get_item_from_api_call(api_base, "api/authn/v2/users?deleted=false&inactive=true&user_name="+ request_encode(username.strip()), username.strip(), "users", "user_name", "user_id", True, verbose, False)
    


def add_field_if_not_blank_or_none(current_content, field_name, field_value):
    if not field_value:
        return current_content
    if field_value == NONE:
        field_value = ''
    if field_name:
        return current_content + f''',
            "{field_name}": "{field_value}"'''
    else:
        return current_content + f''',
            {field_value}'''

def list_allowed_ip_addresses(allowed_ip_addresses):
    if not allowed_ip_addresses:
        return ""
    if allowed_ip_addresses == NONE:
        return '"ip_restricted": false, "allowed_ip_addresses": []'
    all_ip_addresses = allowed_ip_addresses.split(",")
    inner_ip_addresses_list = ""
    for ip_address in all_ip_addresses:
        inner_ip_addresses_list = inner_ip_addresses_list + (""",
            """ if inner_ip_addresses_list else "") + f'"{ip_address.strip()}"'
    if inner_ip_addresses_list:
        return f'''
            "ip_restricted": true, "allowed_ip_addresses": [
                {inner_ip_addresses_list}
            ]'''
    else:
        #this should, realistically, never happen
        return None

def add_permission_based_on_teams(content):
    return content + ''',"permissions":[
      {
         "permission_name":"apiUser"
      }
   ]'''

def modify_user(api_base, user, can_create, generate_credentials, verbose):
    #TODO: add support for creating SAML accounts
    if not user or not user["username"]:
        error_message = "Empty username field found"
        print(error_message)
        return error_message, "", ""

    username = user["username"]
    user_guid = get_user_guid(api_base, username, verbose)

    if not user_guid and not can_create:
        error_message = f"User with name '{username}' not found"
        print(error_message)
        return error_message, "", ""
    
    is_new_user = not user_guid 

    if is_new_user:
        print(f"Creating user: {username}")
    else:
        print(f"Updating user permissions for: {username}")

    if verbose:
        print("Using data:")
        print(user)
    
    if is_new_user:
        url_ending = f"?generate_api_creds={"true" if generate_credentials and user["is_service_account"] else "false"}"
    else:
        url_ending = f"/{user_guid}?partial=true"


    path = f"{api_base}api/authn/v2/users{url_ending}"

    content = f'''"user_name": "{user["username"]}"'''
    if is_new_user and user["is_service_account"]:
        content = add_permission_based_on_teams(content)
    content = add_field_if_not_blank_or_none(content, "active", user["is_active"])
    content = add_field_if_not_blank_or_none(content, "first_name", user["first_name"])
    content = add_field_if_not_blank_or_none(content, "last_name", user["last_name"])
    content = add_field_if_not_blank_or_none(content, "email_address", user["email"])
    content = add_field_if_not_blank_or_none(content, "phone", user["phone"])
    content = add_field_if_not_blank_or_none(content, "title", user["position"])
    content = add_field_if_not_blank_or_none(content, None, list_allowed_ip_addresses(user["restrict_login_ips"]))
    content = add_field_if_not_blank_or_none(content, "login_enabled", str(user["is_login_enabled"] or "").lower())
    content = add_field_if_not_blank_or_none(content, "custom_one", user["custom_1"])
    content = add_field_if_not_blank_or_none(content, "custom_two", user["custom_2"])
    content = add_field_if_not_blank_or_none(content, "custom_three", user["custom_3"])
    content = add_field_if_not_blank_or_none(content, "custom_four", user["custom_4"])
    content = add_field_if_not_blank_or_none(content, "custom_five", user["custom_5"])
    content = add_field_if_not_blank_or_none(content, None, list_roles(user["roles"]))
    content = add_field_if_not_blank_or_none(content, None, list_teams(api_base, user["teams"], user["teams_managed"], verbose))

    request_content=f'''{{
            {content}
        }}'''
    if verbose:
        print(f"Sending {"POST" if is_new_user else "PUT"} request to: {path}")
        print("Request Content:")
        print(request_content)

    if is_new_user:
        response = requests.post(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, json=json.loads(request_content), verify=verify_ssl)
    else:
        response = requests.put(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, json=json.loads(request_content), verify=verify_ssl)

    if verbose:
        print(f"status code {response.status_code}")
        body = response.json()
        if body:
            print(body)
    if response.status_code == 200 or response.status_code == 201:
        if is_new_user:
            print(f"Successfully created {username}.")
        else:
            print(f"Successfully modified user permissions for {username}.")
        if generate_credentials and "api_credentials" in body:
            api_credentials = body["api_credentials"]
            api_id = api_credentials["api_id"]
            api_secret = api_credentials["api_secret"]
        else:
            api_id =""
            api_secret = ""
        return STATUS_SUCCESS, api_id, api_secret
    else:
        body = response.json()
        if (body):
            error_message = f"Operation failed for user {username}: {response.status_code} - {body}"
        else:
            error_message = f"Operation failed for user {username}: {response.status_code}"
        print(error_message)
        return error_message, "", ""
    
def parse_user(excel_sheet, row):
    user = {}
    user["is_service_account"] = excel_sheet.cell(row = row, column = API_SERVICE_ACCOUNT_COLUMN).value
    user["is_active"] = excel_sheet.cell(row = row, column = ACTIVE_COLUMN).value
    user["username"] = excel_sheet.cell(row = row, column = USERNAME_COLUMN).value
    user["first_name"] = excel_sheet.cell(row = row, column = FIRST_NAME_COLUMN).value
    user["last_name"] = excel_sheet.cell(row = row, column = LAST_NAME_COLUMN).value
    user["email"] = excel_sheet.cell(row = row, column = EMAIL_COLUMN).value
    user["phone"] = excel_sheet.cell(row = row, column = PHONE_COLUMN).value
    user["position"] = excel_sheet.cell(row = row, column = POSITION_COLUMN).value
    user["restrict_login_ips"] = excel_sheet.cell(row = row, column = RESTRICT_LOGIN_IPS_COLUMN).value
    user["is_login_enabled"] = excel_sheet.cell(row = row, column = LOGIN_ENABLED_COLUMN).value
    user["custom_1"] = excel_sheet.cell(row = row, column = CUSTOM_1_COLUMN).value
    user["custom_2"] = excel_sheet.cell(row = row, column = CUSTOM_2_COLUMN).value
    user["custom_3"] = excel_sheet.cell(row = row, column = CUSTOM_3_COLUMN).value
    user["custom_4"] = excel_sheet.cell(row = row, column = CUSTOM_4_COLUMN).value
    user["custom_5"] = excel_sheet.cell(row = row, column = CUSTOM_5_COLUMN).value
    user["teams"] = excel_sheet.cell(row = row, column = TEAMS_COLUMN).value
    user["roles"] = excel_sheet.cell(row = row, column = ROLES_COLUMN).value
    user["teams_managed"] = excel_sheet.cell(row = row, column = TEAMS_MANAGED_COLUMN).value

    return user
    

def modify_all_users(api_base, file_name, can_create, generate_credentials, verbose):
    global failed_attempts
    excel_file = openpyxl.load_workbook(file_name)
    excel_sheet = excel_file.active    
    try:
        for row in range(FIRST_ROW, excel_sheet.max_row+1):
            failed_attempts = 0
            status=excel_sheet.cell(row = row, column = STATUS_COLUMN).value
            if (status == STATUS_SUCCESS):
                print(f"Skipping row {row-FIRST_ROW+1} as it was already done (physical row: {row})")
            else:
                try:
                    print(f"Importing row {row-FIRST_ROW+1}/{excel_sheet.max_row-FIRST_ROW+1} (physical row: {row}):")
                    status, api_id, api_secret = modify_user(api_base, 
                                                 parse_user(excel_sheet, row),
                                                 can_create, 
                                                 generate_credentials,
                                                 verbose)
                    print(f"Finished importing row {row-FIRST_ROW+1}/{excel_sheet.max_row-FIRST_ROW+1} (physical row: {row})")
                    print("---------------------------------------------------------------------------")
                except (NoExactMatchFoundException, UnableToCreateTeamException, NoResultFoundException) as e:
                    status= e.get_message()
                    api_id = ""
                    api_secret = ""
                excel_sheet.cell(row = row, column = STATUS_COLUMN).value=status
                excel_sheet.cell(row = row, column = API_ID_COLUMN).value=api_id
                excel_sheet.cell(row = row, column = API_SECRET_COLUMN).value=api_secret
    finally:
        excel_file.save(filename=file_name)

def get_api_base():
    api_key_id, api_key_secret = get_credentials()
    api_base = "https://api.veracode.{instance}/"
    if api_key_id.startswith("vera01"):
        return api_base.replace("{instance}", "eu", 1)
    else:
        return api_base.replace("{instance}", "com", 1)

def main(argv):
    """Allows for bulk creation or modifying user and permissions"""
    global failed_attempts
    global verify_ssl
    excel_file = None
    try:
        verbose = False
        can_create = False
        generate_credentials = False
        file_name = ''

        opts, args = getopt.getopt(argv, "hdcgfv:", ["file_name=","verify_ssl"])
        for opt, arg in opts:
            if opt == '-h':
                print_help()
            if opt == '-d':
                verbose = True
            if opt == '-c':
                can_create = True
            if opt == '-g':
                generate_credentials = True
            if opt in ('-v', '--verify_ssl'):
                verify_ssl=arg.strip().lower() == "true"
            if opt in ('-f', '--file_name'):
                file_name=arg

        api_base = get_api_base()
        if file_name:
            modify_all_users(api_base, file_name, can_create, generate_credentials, verbose)
        else:
            print_help()
    except requests.RequestException as e:
        print("An error occurred!")
        print(e)
        sys.exit(1)
    finally:
        if excel_file:
            excel_file.save(filename=file_name)


if __name__ == "__main__":
    main(sys.argv[1:])
